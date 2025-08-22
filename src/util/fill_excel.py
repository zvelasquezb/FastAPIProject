from openpyxl import Workbook, load_workbook
import shutil
import os

def escribir_en_excel(ruta_archivo, nombre_hoja, celda, valor):
    """
    Escribe un valor en una celda específica de una hoja de cálculo en un archivo de Excel.

    Args:
        ruta_archivo (str): La ruta completa del archivo de Excel (ej. 'C:/mis_datos/reporte.xlsx').
        nombre_hoja (str): El nombre de la hoja de cálculo.
        celda (str): La celda de destino (ej. 'A1', 'B5').
        valor: El valor que se va a escribir en la celda.
    """
    try:
        # Cargar el libro de trabajo si existe. Si no, crear uno nuevo.
        try:
            libro = load_workbook(ruta_archivo)
        except FileNotFoundError:
            print(f"El archivo '{ruta_archivo}' no existe. Creando uno nuevo.")
            libro = Workbook()
            # La primera hoja se crea automáticamente, le ponemos el nombre especificado
            libro.active.title = nombre_hoja
        
        # Seleccionar la hoja. Si no existe, crearla.
        if nombre_hoja in libro.sheetnames:
            hoja = libro[nombre_hoja]
        else:
            print(f"La hoja '{nombre_hoja}' no existe. Creando una nueva.")
            hoja = libro.create_sheet(nombre_hoja)
        
        # Escribir el valor en la celda especificada
        hoja[celda] = valor
        
        # Guardar los cambios en el archivo
        libro.save(ruta_archivo)
        print(f"Valor '{valor}' escrito correctamente en la celda {celda} de la hoja '{nombre_hoja}'.")
    
    except Exception as e:
        print(f"Ocurrió un error al intentar escribir en el archivo: {e}")

# --- Ejemplo de uso ---
# Reemplaza la ruta, hoja y celda con tus propios datos
# escribir_en_excel('C:/proyectos/mi_reporte.xlsx', 'Ventas Anuales', 'A1', 'Total de Ventas')
# escribir_en_excel('C:/proyectos/mi_reporte.xlsx', 'Ventas Anuales', 'B2', 125000)
# escribir_en_excel('C:/proyectos/mi_reporte.xlsx', 'Inventario', 'C3', 'Producto XYZ')


def es_celda_vacia(ruta_archivo, nombre_hoja, celda):
    """
    Verifica si una celda específica en un archivo de Excel está vacía.

    Args:
        ruta_archivo (str): La ruta completa del archivo de Excel.
        nombre_hoja (str): El nombre de la hoja de cálculo.
        celda (str): La celda a verificar (ej. 'A1', 'B5').

    Returns:
        bool: True si la celda está vacía, False si contiene un valor.
    """
    try:
        # Cargar el libro de trabajo
        libro = load_workbook(ruta_archivo)
        
        # Seleccionar la hoja
        if nombre_hoja in libro.sheetnames:
            hoja = libro[nombre_hoja]
        else:
            print(f"Error: La hoja '{nombre_hoja}' no se encontró en el archivo.")
            return True # Consideramos que no existe, por lo que está "vacía"

        # Obtener el valor de la celda
        valor_celda = hoja[celda].value
        
        # Comprobar si el valor es None o una cadena vacía después de limpiar espacios en blanco
        return valor_celda is None or (isinstance(valor_celda, str) and valor_celda.strip() == "")

    except FileNotFoundError:
        print(f"Error: El archivo '{ruta_archivo}' no se encontró.")
        return True # El archivo no existe, por lo que la celda está "vacía"
    except Exception as e:
        print(f"Ocurrió un error inesperado: {e}")
        return True

# --- Ejemplo de uso ---
# Supongamos que tienes un archivo "datos.xlsx" con la hoja "Hoja1"
# y la celda A1 tiene "Hola" y la celda B2 está vacía.
#
# print(es_celda_vacia("datos.xlsx", "Hoja1", "A1")) # Debería imprimir False
# print(es_celda_vacia("datos.xlsx", "Hoja1", "B2")) # Debería imprimir True



def copiar_y_renombrar_excel(ruta_origen, ruta_destino_con_nuevo_nombre):
    """
    Copia un archivo de Excel completo a una nueva ubicación con un nuevo nombre.

    Args:
        ruta_origen (str): La ruta completa del archivo de Excel original.
        ruta_destino_con_nuevo_nombre (str): La ruta completa del nuevo archivo,
                                            incluyendo el nuevo nombre.
    """
    try:
        # Asegurarse de que el archivo de origen existe
        if not os.path.exists(ruta_origen):
            print(f"Error: El archivo de origen '{ruta_origen}' no se encontró.")
            return

        # Copiar el archivo. shutil.copy() copia el contenido y metadata.
        # Si la ruta de destino incluye un nuevo nombre, el archivo copiado
        # se guarda con ese nombre.
        shutil.copy(ruta_origen, ruta_destino_con_nuevo_nombre)
        
        print(f"Archivo copiado exitosamente de:\n'{ruta_origen}'\na:\n'{ruta_destino_con_nuevo_nombre}'")

    except Exception as e:
        print(f"Ocurrió un error al copiar el archivo: {e}")

# --- Ejemplo de uso ---
# Reemplaza estas rutas con las tuyas
#ruta_original = "C:\\MisDocumentos\\reporte_ventas.xlsx"
#ruta_nueva = "C:\\MisDocumentos\\reporte_ventas_2025.xlsx"

# Llama a la función para copiar el archivo
#copiar_y_renombrar_excel(ruta_original, ruta_nueva)

def borrar_archivo_si_existe(ruta_archivo):
    """
    Intenta borrar un archivo en la ruta especificada.

    Args:
        ruta_archivo (str): La ruta completa del archivo que se va a borrar.
    """
    try:
        # 1. Verificar si el archivo existe
        if os.path.exists(ruta_archivo):
            # 2. Si existe, intenta borrarlo
            os.remove(ruta_archivo)
            print(f"El archivo '{ruta_archivo}' ha sido borrado exitosamente.")
        else:
            # 3. Si no existe, informa sin generar un error
            print(f"El archivo '{ruta_archivo}' no existe. No es necesario borrarlo.")
    except Exception as e:
        # 4. Manejar cualquier otro error (ej. permisos)
        print(f"Ocurrió un error al intentar borrar el archivo: {e}")

# --- Ejemplo de uso ---
# Reemplaza la ruta con la de tu archivo
# archivo_a_borrar = "C:\\MisDocumentos\\archivo_temporal.txt"
# borrar_archivo_si_existe(archivo_a_borrar)


def fill_excel(results):
    row=3
    saldo_inicial=["I",row]
    movimiento_de_ingresos_en_pesos=["J",row]
    movimiento_de_egresos_en_pesos=["K",row]
    saldo_en_pesos_al_final_de_mes_segun_tesoreria=["L",row]
    valor_de_movimiento_maximo_en_el_mes_en_pesos=["M",row]

    borrar_archivo_si_existe(r"src\template_excel\Inf SIVICOF_PLANTILLA.xlsx")
    copiar_y_renombrar_excel(r"src\template_excel\1Inf SIVICOF_PLANTILLA.xlsx",r"src\template_excel\Inf SIVICOF_PLANTILLA.xlsx")
    for length in range(len(results)):
        if(es_celda_vacia(r"src\template_excel\Inf SIVICOF_PLANTILLA.xlsx","14233 CB-0115  INFORME SOBR...",str(saldo_inicial[0]+str(saldo_inicial[1])))):
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(saldo_inicial[0]+str(saldo_inicial[1])), results[length]['data']['SALDO ANTERIOR'])
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(movimiento_de_ingresos_en_pesos[0]+str(movimiento_de_ingresos_en_pesos[1])), results[length]['data']['TOTAL ABONOS'])
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(movimiento_de_egresos_en_pesos[0]+str(movimiento_de_egresos_en_pesos[1])), results[length]['data']['TOTAL CARGOS'])
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(saldo_en_pesos_al_final_de_mes_segun_tesoreria[0]+str(saldo_en_pesos_al_final_de_mes_segun_tesoreria[1])), results[length]['data']['SALDO ACTUAL'])
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(valor_de_movimiento_maximo_en_el_mes_en_pesos[0]+str(valor_de_movimiento_maximo_en_el_mes_en_pesos[1])), results[length]['data']['Valor de movimiento maximo en el mes en pesos'])
            saldo_inicial[1]+=1
            movimiento_de_ingresos_en_pesos[1]+=1
            movimiento_de_egresos_en_pesos[1]+=1
            saldo_en_pesos_al_final_de_mes_segun_tesoreria[1]+=1
            valor_de_movimiento_maximo_en_el_mes_en_pesos[1]+=1
        else:
            saldo_inicial[1]+=1
            movimiento_de_ingresos_en_pesos[1]+=1
            movimiento_de_egresos_en_pesos[1]+=1
            saldo_en_pesos_al_final_de_mes_segun_tesoreria[1]+=1
            valor_de_movimiento_maximo_en_el_mes_en_pesos[1]+=1
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(saldo_inicial[0]+str(saldo_inicial[1])), results[length]['data']['SALDO ANTERIOR'])
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(movimiento_de_ingresos_en_pesos[0]+str(movimiento_de_ingresos_en_pesos[1])), results[length]['data']['TOTAL ABONOS'])
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(movimiento_de_egresos_en_pesos[0]+str(movimiento_de_egresos_en_pesos[1])), results[length]['data']['TOTAL CARGOS'])
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(saldo_en_pesos_al_final_de_mes_segun_tesoreria[0]+str(saldo_en_pesos_al_final_de_mes_segun_tesoreria[1])), results[length]['data']['SALDO ACTUAL'])
            escribir_en_excel(r'src\template_excel\Inf SIVICOF_PLANTILLA.xlsx',"14233 CB-0115  INFORME SOBR...", str(valor_de_movimiento_maximo_en_el_mes_en_pesos[0]+str(valor_de_movimiento_maximo_en_el_mes_en_pesos[1])), results[length]['data']['Valor de movimiento maximo en el mes en pesos'])
